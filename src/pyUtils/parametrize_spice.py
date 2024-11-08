#!/usr/bin/python3

# This file is part of the EmCAD program which constitutes the client
# side of an electromagnetic modeler delivered as a cloud based service.
# 
# Copyright (C) 2015-2020  Walter Steffe
# 
# This program is free software; you can redistribute it and/or
# modify it under the terms of the GNU General Public License
# as published by the Free Software Foundation; either version 2
# of the License, or (at your option) any later version.
# 
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
# 
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.


import sys,os,time
from PyQt5 import QtWidgets, QtCore
import numpy, math
import scipy.optimize as optim
import json
import openpyxl
from cvxopt import solvers
from cvxopt import matrix


from enum import Enum
class ErrorCode(Enum):
     singularJaco = 1


argv=sys.argv
argl=len(argv)

check=False
for i,arg in enumerate(argv):
     if arg=="-check":
          argv.pop(i)
          check=True
          break

def touch(fname, times=None):
    with open(fname, 'a'):
        os.utime(fname, times)


def goodBroydenUpdate(jact, derr, dx):
      nx=len(dx)
      ne=len(derr)
      dxt=numpy.reshape(dx,(1,nx))
      derrt=numpy.reshape(derr,(1,ne))
      derrt=derrt-numpy.dot(dxt,jact)
      dxv=numpy.reshape(dx,(nx,1))
      djact=numpy.dot(dxv,derrt)/numpy.dot(dx,dx)
      jact=jact+djact
      return jact


class Tuner(QtCore.QObject):

    def __init__(self):    
        QtCore.QObject.__init__(self)

        argv=sys.argv
        argl=len(argv)

        self.var_fname=''
        for i,arg in enumerate(argv):
            if arg=="-varfile":
                argv.pop(i)
                self.var_fname=argv.pop(i)
                break

        self.method=0
        for i,arg in enumerate(argv):
            if arg=="-method":
                argv.pop(i)
                self.method=int(argv.pop(i))
                break

        self.maxiter=1000
        for i,arg in enumerate(argv):
            if arg=="-maxiter":
                argv.pop(i)
                self.maxiter=int(argv.pop(i))
                break

        self.xtol=1.e-4
        for i,arg in enumerate(argv):
            if arg=="-xtol":
                argv.pop(i)
                self.xtol=float(argv.pop(i))
                break

        self.trustd=1
        for i,arg in enumerate(argv):
            if arg=="-trustd":
                argv.pop(i)
                self.trustd=float(argv.pop(i))
                break


        self.x_number_of_decimals=-int(round(numpy.log10(self.xtol)))


        self.recomputeerror=False
        for i,arg in enumerate(argv):
            if arg=="-recomputeerror":
                argv.pop(i)
                self.recomputeerror=True
                break

        self.recomputejaco=False
        for i,arg in enumerate(argv):
            if arg=="-recomputejaco":
                argv.pop(i)
                self.recomputejaco=True
                break

        self.ideal_cktname=''
        for i,arg in enumerate(argv):
            if arg=="-ideal_cktname":
                argv.pop(i)
                self.ideal_cktname=argv.pop(i)
                break

        self.mapped_cktname=''
        for i,arg in enumerate(argv):
            if arg=="-mapped_cktname":
                argv.pop(i)
                self.mapped_cktname=argv.pop(i)
                break

        self.check=False
        for i,arg in enumerate(argv):
            if arg=="-check":
                argv.pop(i)
                self.check=True
                break

        self.ideal_fname=self.ideal_cktname+'_xpar.txt'
        self.mapped_fname=self.mapped_cktname+'_xpar.txt'
        self.mapped_dfname=self.mapped_cktname+'.done'

        var_fname_split=self.var_fname.split(".")
        self.var_ott_fname=var_fname_split[0]+"_ott."+var_fname_split[1]





    signal_status0 = QtCore.pyqtSignal(['QString'])
    def status0(self, txt):
        self.signal_status0.emit(txt)

    signal_status1 = QtCore.pyqtSignal(['QString'])
    def status1(self, txt):
        self.signal_status1.emit(txt)

    signal_status2 = QtCore.pyqtSignal(['QString'])
    def status2(self, txt):
        self.signal_status2.emit(txt)

    def on_break(self):
        self.cntrl_break=True

    def check_files(self):
        self.varnames, self.xstart= self.readTable(self.var_fname, '\t')
        self.ideal_par_names, self.ideal_par_values=self.readTable(self.ideal_fname)
        nvar=len(self.xstart)
        npar=len(self.ideal_par_names)
#        import pdb
#        pdb.set_trace()
        if nvar==0:
            return 1
        elif npar==0:
            return 2
        else:
            return 0

    def init(self):
        self.varnames, self.xstart= self.readTable(self.var_fname, '\t')
        self.ideal_par_names, self.ideal_par_values=self.readTable(self.ideal_fname)
        self.ideal_par=self.make_dict(self.ideal_par_names, self.ideal_par_values)

        self.EXITCODE=0
        self.cntrl_break=False

        self.evaluatedF={}
        self.evaluatedJaco={}

#        if self.recomputeerror or not self.updatedMap():
        if self.recomputeerror:
            self.initial_err=numpy.array([])
        else:
            self.touchMap()
            self.initial_err=numpy.array(self.readError())
            npar=len(self.ideal_par_names)
            if len(self.initial_err)==npar:
                key=self.x_key(self.xstart)
                self.evaluatedF[key]=self.initial_err
    
        if self.recomputejaco:
            self.initial_jact=numpy.array([])
        else:
            self.initial_jact=numpy.array(self.readJaco())


    def x_key(self,x):
        key= [ '{:10.{precision}f}'.format(x[i],precision=self.x_number_of_decimals) for i in range(len(x))]
        return tuple(key)

    def x_val(self,key):
        x= [float(key[i]) for i in range(len(key))]
        return x

    def readJaco(self):
        if(not os.path.isfile('jacobian.json')):
            return []
        with open('jacobian.json') as data_file:    
            jact = json.load(data_file)
        return jact

    def writeJaco(self,jact):
        with open('jacobian.json', 'w') as data_file:
            json.dump(jact, data_file)

    def readError(self):
        if(not os.path.isfile('error.json')):
            return []
        with open('error.json') as data_file:    
            err = json.load(data_file)
        return err

    def writeError(self,fname,err):
        with open(fname, 'w') as data_file:
            json.dump(err, data_file)

    def readSpCircuit(self,fname):
        names=[]
        values=[]
#        import pdb
#        pdb.set_trace()
        with open(fname) as data_file:
            for line in data_file:
                lsplit=line.rstrip('\r\n').split()
                name=lsplit[0]
                val=lsplit[-1]
                if not name.startswith(".") and not name.startswith("+")
                   names.append(name)
                   values.append(float(val))
        return names, numpy.array(values)


    def writeTable(self,fname,names, values, sep=":", fname0="-"):
        if fname.endswith('.txt') or fname.endswith('.csv'):
          f=open(fname, 'w')
          for i in range(len(names)):
            f.write(names[i]+sep)
            f.write('{:10.{precision}f}'.format(values[i],precision=self.x_number_of_decimals) +"\r\n")
#           f.write(str(values[i])+"\r\n")
          f.close()
        elif fname.endswith('.xlsx'):
            if fname0=="-":
              wb=openpyxl.load_workbook(fname)
            else:
              wb=openpyxl.load_workbook(fname0)
            sheet =wb.get_sheet_by_name('Sheet1')
            i=0
            for rowNum in range(2, sheet.max_row+1): # skip the first row
                parType = sheet.cell(row=rowNum, column=1).value
                if parType=='Variables':
                    sheet.cell(row=rowNum, column=3).value ='{:10.{precision}f}'.format(values[i],precision=self.x_number_of_decimals)
                    sheet.cell(row=rowNum, column=4).value ='{:10.{precision}f}'.format(values[i],precision=self.x_number_of_decimals)
                    i=i+1
            wb.save(fname)
            wb.close()

    def make_dict(self,names, values):
        d={}
        for i in range(len(names)):
            d[names[i]]=values[i]
        return d


    def clearError(self,err):
        l=len(err)
        err= [0 for i in range(l)]
        return numpy.array(err)


    def singularJaco(self,dJ):
        tol=1.e-10
        nvar=len(dJ)
        npar=len(self.ideal_par_names)
        n=min(npar,nvar)
        if n==0:
            return False
        jac=numpy.transpose(numpy.array(dJ))
        r=numpy.linalg.qr(jac,mode='r')
        if abs(r[n-1,n-1])>tol:
            return False
        else:
            return True


    def updatedMap(self):
        if not os.path.exists(self.mapped_dfname):
           return False
        var_t=os.path.getmtime(self.var_fname)
        map_t=os.path.getmtime(self.mapped_dfname)
        return map_t>var_t

    def touchMap(self):
        touch(self.mapped_fname)


    def getError(self):
#       let user to perform cad update and electromagnetic analysis
        err=[]
        while not self.updatedMap() and not self.cntrl_break:
            self.status1('Waiting for simulation')
            time.sleep(1)  # Delay for 1 second
            if self.cntrl_break:
                return numpy.array(err)
        self.mapped_par_names, self.mapped_par_values=self.readTable(self.mapped_fname)
        self.mapped_par=self.make_dict(self.mapped_par_names, self.mapped_par_values)
        for k in self.ideal_par_names:
            if k in self.mapped_par: 
                e=float(self.mapped_par[k])-float(self.ideal_par[k])
            else:
                e=-float(self.ideal_par[k])
            err.append(e)
        return numpy.array(err)

# error function:
    def error(self,x, store=True):
        key=self.x_key(x)
        if key in self.evaluatedF:
            return self.evaluatedF[key]
        self.writeTable(self.var_fname,self.varnames,x,'\t')
#   get ana output
        if self.cntrl_break:
            err=self.clearError(err)
            return err
        else:
            err=self.getError()
        if(store and len(err)>0):
            self.evaluatedF[key]=err
        return err

    def evalJacoT(self,x):
        dx=self.trustd
        current_err=self.error(x)
        dJt=self.initial_jact.tolist()
        n0=len(dJt)
        n=len(x)
        for i in range(n0,n):
            x1=numpy.copy(x)
            x1[i]=x[i]+dx
            self.status0('Evaluating Parameter Derivatives vs. x'+'%d'%(i+1))
            err1=self.error(x1,False)
            if self.cntrl_break:
                break 
            derr=(err1-current_err)/dx
            txt2='\nParameter Derivatives vs. x'+'%d'%(i+1)+':\n'
            for j in range(len(self.ideal_par_names)):
                txt2=txt2+'D('+self.ideal_par_names[j]+') = '+'%12.8f'%derr[j]+'\n'
            self.status2(txt2)
            dJt.append(derr.tolist())
            if self.singularJaco(dJt):
                self.cntrl_break=True
                self.EXITCODE=ErrorCode.singularJaco
                self.status0('Tuner exited with error')
                self.status1('Parameter derivatives are not linearly independent')
                return dJt
            self.writeJaco(dJt)
            if self.cntrl_break:
                break 
        return numpy.array(dJt)


    def jaco(self,x):
        key=self.x_key(x)
        if key in self.evaluatedJaco:
            return self.evaluatedJaco[key]
        else:
            dJt=self.evalJacoT(x)
            self.evaluatedJaco[key]=dJt
            return dJt


    def tune0(self):
#        import pdb
#        pdb.set_trace()
        err=numpy.copy(self.initial_err)
        self.xcurrent=numpy.copy(self.xstart)
        if len(err)==0:
            self.status0('Evalutating initial error')
            err=self.error(self.xstart)
            if self.cntrl_break:
                return numpy.array([]) 
            self.writeError('error.json',err.tolist())

        nvar=len(self.xstart)
        npar=len(self.ideal_par_names)

        err_norm=numpy.linalg.norm(err)
        txt2='\nInitial Error Norm = '+'%12.8f'%err_norm+'\n'
        txt2=txt2+'Initial Errors: \n'
        for j in range(npar):
            txt2=txt2+'Error('+self.ideal_par_names[j]+') = '+'%12.8f'%err[j]+'\n'
        self.status2(txt2)

        jact=numpy.copy(self.initial_jact)
        if len(jact)<nvar:
            jact=self.evalJacoT(self.xstart)
            if self.cntrl_break:
                return self.xstart
                self.status0('Tuner exited')
                self.status1('Break entered')
#        invjact=numpy.linalg.inv(jact)
#        invjact=numpy.transpose(invjact)

        x=numpy.copy(self.xstart)
        self.xott=numpy.copy(self.xstart)
        err_norm_min=err_norm
        for i in range(self.maxiter):
            grad=matrix(jact.dot(err))
            hess=matrix(jact.dot(numpy.transpose(jact)))
            G=numpy.zeros((2*nvar,nvar))
            numpy.fill_diagonal(G[:nvar,:nvar],1)
            numpy.fill_diagonal(G[nvar:2*nvar,:nvar],-1)
            G=matrix(G)
            h=matrix([self.trustd for i in range(2*nvar)])
            sol=solvers.qp(hess,grad,G,h)
            dx=(numpy.array(sol['x'])).flatten()
            dx_norm=numpy.linalg.norm(dx,ord=numpy.inf)
            if(dx_norm<self.xtol):
                self.status0('Tuner task terminated')
                self.status1('Variable displacement is less than xtol')
                return
            x=x+dx
            self.status0('Iteration '+'%d'%(i+1)+' : evalutating Error')
            if(i>0):
                txt2='\nError Norm at Iteration '+'%d'%i+' = '+'%12.8f'%err_norm+'\n'
                txt2=txt2+'Errors at Iteration '+'%d'%i+':\n'
                for j in range(npar):
                    txt2=txt2+'Error('+self.ideal_par_names[j]+') = '+'%12.8f'%err[j]+'\n'
            txt2=txt2+'Best Error Norm='+'%12.8f'%err_norm_min+'\n'
            txt2=txt2+'\nDx Norm of Iteration '+'%d'%(i+1)+' = '+'%12.8f'%dx_norm+'\n'
            self.status2(txt2)
            err0=numpy.copy(err)
            err=self.error(x)
            if self.cntrl_break:
                self.status0('Tuner exited')
                self.status1('Break entered')
                return
            #xcurrent holds last evalueted parameters
            self.writeError('error.json',err.tolist())
            self.xcurrent=numpy.copy(x)  
            if dx_norm>self.trustd/2:
                derr=err-err0
                jact=goodBroydenUpdate(jact, derr, dx)
                dJt=jact.tolist()
                self.writeJaco(dJt)
            err_norm=numpy.linalg.norm(err)
            if(err_norm<err_norm_min):
                 err_norm_min=err_norm
                 self.xott=numpy.copy(x)
                 self.writeTable(self.var_ott_fname,self.varnames,self.xott,"\t", self.var_fname)
                 self.writeError('error_ott.json',err.tolist())
        self.status0('Tuner task terminated')
        self.status1('Max iter reached')


    def tune1(self, xstart, xtol=1.e-3, scale=None, niter=10):
        if scale==None: scale= [1 for i in range(len(xstart))]
        self.xott=optim.fsolve(self.error, xstart, fprime=self.jaco, col_deriv=True, xtol=xtol, maxfev=niter, factor=0.1, diag=scale)

    def tune2(self, xstart, xtol=1.e-3, scale=None, niter=10):
        if scale==None: scale= [1 for i in range(len(xstart))]
        self.xott=optim.broyden1(self.error, xstart, maxiter=iter, f_tol=1.e-7, x_tol=xtol)

    def tune3(self, xstart, xtol=1.e-3, scale=None, niter=10):
        self.xott=optim.fsolve(self.error, xstart, epsfcn=20*xtol, xtol=xtol, maxfev=niter, factor=0.1)

    def tune4(self, xstart, xtol=1.e-3, scale=None, niter=10):
        if scale==None: scale= [1 for i in range(len(xstart))]
        self.xott=optim.leastsq(self.error, xstart, xtol=xtol, maxfev=niter, epsfcn=10*xtol, factor=0.1, diag=scale)

    def tune5(self, xstart, xtol=1.e-3, scale=None, niter=10):
        if scale==None: scale= [1 for i in range(len(xstart))]
        options={'eps': 1.e-2, 'maxfev': niter, 'xtol' : xtol, 'factor' : 0.1, 'diag' : scale}
        self.xott=optim.root(error, xstart, method='hybr', tol=1.e-6)

    def tune6(self, xstart, xtol=1.e-3, scale=None, niter=10):
        if scale==None: scale= [1 for i in range(len(xstart))]
        options={'maxfev': niter, 'xtol' : xtol, 'factor' : 0.1, 'col_deriv' : True, 'diag' : scale}
        sol=optim.root(self.error, xstart, method='hybr', jac=self.jaco, tol=1.e-6)
        self.xott=list(sol.x)

    def tune(self):
        # Note: This is never called directly. It is called by Qt once the
        # thread environment has been set up.
        if self.method==0:
            self.init()
            self.tune0()
        if self.method==1:
            self.init()
            self.tune1(self.xstart, self.xtol, niter=self.maxiter)
        self.writeTable(self.var_fname,self.varnames,self.xcurrent,"\t")
        return self.EXITCODE



class Worker(QtCore.QThread):
    def __init__(self, parent = None):    
        QtCore.QThread.__init__(self, parent = None)
        self.exiting = False
        self.tuner=Tuner()
        
    signal_exiting = QtCore.pyqtSignal()
    def emit_exiting(self):
        self.signal_exiting.emit()

    def run(self):
#        import pdb
#        pdb.set_trace()
        self.tuner.init()
        self.tuner.tune()
        self.emit_exiting()
        while True:
            time.sleep(10)

#    def __del__(self):
#        self.wait()


class CntrlGui(QtWidgets.QMessageBox):
    def __init__(self):
        QtWidgets.QMessageBox.__init__(self)
        global cntrl_break, status
        self.setSizeGripEnabled(True)


    def event(self, e):
        result = QtWidgets.QMessageBox.event(self, e)

        self.setMinimumHeight(0)
        self.setMaximumHeight(16777215)
        self.setMinimumWidth(500)
        self.setMaximumWidth(16777215)
        self.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)

        textEdit = self.findChild(QtWidgets.QTextEdit)
        if textEdit != None :
            textEdit.setMinimumHeight(0)
            textEdit.setMaximumHeight(16777215)
            textEdit.setMinimumWidth(500)
            textEdit.setMaximumWidth(16777215)
            textEdit.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        return result



class Window(QtWidgets.QWidget):

    signal_break = QtCore.pyqtSignal()
    
    def __init__(self, parent = None):
    
        QtWidgets.QWidget.__init__(self, parent)
        
        self.thread = Worker()

        self.thread.tuner.signal_status0.connect(self.status0)
        self.thread.tuner.signal_status1.connect(self.status1)
        self.thread.tuner.signal_status2.connect(self.status2)
        self.thread.signal_exiting.connect(self.on_thtread_exiting)
        self.setWindowTitle(self.tr("Filter Tuning Wizard"))
        self.text0 = QtWidgets.QLabel(self.tr(""))
        self.text1 = QtWidgets.QLabel(self.tr(""))
        self.text2 = QtWidgets.QTextEdit(self.tr(""))
        self.text2.setReadOnly(True)
        self.break_btn =QtWidgets.QPushButton('Break')
        self.close_btn =QtWidgets.QPushButton('Close')
        self.close_btn.setEnabled(False);
        self.break_btn.clicked.connect(self.emit_break)
        self.signal_break.connect(self.thread.tuner.on_break)
        self.close_btn.clicked.connect(app.exit)
        self.mainWidget = QtWidgets.QWidget(self)
        self.mainLayout = QtWidgets.QVBoxLayout(self.mainWidget)
        self.hbox = QtWidgets.QHBoxLayout()
        self.hbox.addWidget(self.break_btn)
        self.hbox.addWidget(self.close_btn)
        self.hbox.addStretch(1)
        self.mainLayout.addWidget(self.text0)
        self.mainLayout.addWidget(self.text1)
        self.mainLayout.addLayout(self.hbox)
        self.mainLayout.addWidget(self.text2)
        self.setLayout(self.mainLayout)
        
        self.setMinimumHeight(0)
        self.setMaximumHeight(16777215)
        self.setMinimumWidth(500)
        self.setMaximumWidth(16777215)
        self.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)

        textEdit = self.findChild(QtWidgets.QTextEdit)


        self.thread.start()

    def emit_break(self):
        self.signal_break.emit()

    def status0(self, txt):
        self.text0.setText(txt)

    def status1(self, txt):
        self.text1.setText(txt)

    def status2(self, txt):
        self.text2.append(txt)

    def on_thtread_exiting(self):
        self.close_btn.setEnabled(True);


if __name__ == "__main__":
    if check:
       tuner=Tuner()
       e=tuner.check_files()
       if(e==0):
         sys.stdout.write("exitCode=0"+ '\n')
         sys.exit(e)
       else:
         sys.stdout.write("exitCode=1"+ '\n')
         sys.exit(e)
    else:
       app = QtWidgets.QApplication(sys.argv)
       window = Window()
       window.show()
       sys.exit(app.exec_())




